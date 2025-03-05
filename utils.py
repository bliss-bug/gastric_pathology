import os
import random
import openpyxl
import torch
from sklearn.model_selection import KFold
from collections import defaultdict


def load_data(data_path, label_path, fold):
    workbook = openpyxl.load_workbook(label_path)
    sheet = workbook.active
    rows = sheet.iter_rows()

    labels = {}
    for i, row in enumerate(rows):
        if i > 0:
            id, label = str(row[1].value), row[2].value
            labels[id] = label

    if isinstance(data_path, str):
        feats_path = os.listdir(data_path)
        feats_path = [os.path.join(data_path, feat_path) for feat_path in feats_path]
    elif isinstance(data_path, list):
        feats_path = []
        for path in data_path:
            temp_path = [os.path.join(path, feat_path) for feat_path in os.listdir(path)]
            feats_path.extend(temp_path)

    random.shuffle(feats_path)
    
    n = len(feats_path)
    print(n)

    kf = KFold(n_splits=5)
    split = list(kf.split(feats_path))[(fold-1)%5]
    
    train_path, test_path = [feats_path[i] for i in split[0]], [feats_path[i] for i in split[1]]
    print(f'fold{fold} test data: {test_path}')

    random.shuffle(train_path)
    div = int(len(train_path)*0.875)
    train_path, val_path = train_path[:div], train_path[div:]
    #print(val_path)
    '''
    p, q, w = [], [], []

    for feat_path in feats_path:
        id = feat_path.split('/')[-1].split('.')[0]
        label = labels[id] if id in labels else -1
        if label>=0:
            p.append(id)
        else:
            q.append(id)

    for c in labels.keys():
        if c not in p:
            w.append(c)

    print(q, w)
    '''
    return train_path, val_path, test_path, labels



def load_trainval_data(data_path, label_path):
    workbook = openpyxl.load_workbook(label_path)
    sheet = workbook.active
    rows = sheet.iter_rows()

    labels = {}
    for i, row in enumerate(rows):
        if i > 0:
            id, label = str(row[1].value), row[2].value
            labels[id] = label

    if isinstance(data_path, str):
        feats_path = os.listdir(data_path)
        feats_path = [os.path.join(data_path, feat_path) for feat_path in feats_path]
    elif isinstance(data_path, list):
        feats_path = []
        for path in data_path:
            temp_path = [os.path.join(path, feat_path) for feat_path in os.listdir(path)]
            feats_path.extend(temp_path)

    #random.shuffle(feats_path)
    mp = defaultdict(list)
    for feat_path in feats_path:
        id = feat_path.split('/')[-1].split('.')[0].split('-')[0].split('H')[0]
        mp[id].append(feat_path)
    parents = list(mp.keys())
    random.shuffle(parents)

    div = int(len(parents)*0.9)
    train_path, val_path = [], []
    for i in range(div):
        train_path.extend(mp[parents[i]])
    for i in range(div, len(parents)):
        val_path.extend(mp[parents[i]])
    
    return train_path, val_path, labels



def load_test_data(data_path, label_path):
    workbook = openpyxl.load_workbook(label_path)
    sheet = workbook.active
    rows = sheet.iter_rows()

    labels = {}
    for i, row in enumerate(rows):
        if i > 0:
            id, label = str(row[1].value), row[2].value
            labels[id] = label

    feats_path = os.listdir(data_path)
    feats_path = [os.path.join(data_path, feat_path) for feat_path in feats_path]
    
    return feats_path, labels



def get_cam_1d(classifier, features):
    tweight = list(classifier.parameters())[-2]
    cam_maps = torch.einsum('bgf,cf->bcg', [features, tweight])
    return cam_maps


if __name__ == '__main__':
    random.seed(42)
    #load_data(['WSI/features/uni_features', 'WSI/features2/uni_features', 'WSI/features3/uni_features'], 'labels/NDPI_labels.xlsx', 5)
    #load_test_data('WSI/features/gigapath_features', 'labels/NDPI_labels.xlsx')
    load_trainval_data(['WSI/features_in/gigapath_features'], 'labels/all_labels.xlsx')